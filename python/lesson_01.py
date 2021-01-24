# print("全程班78期班花柠檬")
# print("全程班77期班草冬瓜")
# print("全程班78期班草六六")
print(type(12))
# print(type(3.1415926))
# """print(isinstance(12,int))
# print("全程班78期班花是'依依'")
# print('''----圆圆的基本信息----
#     name:圆圆
#     gender：female
#     hobby：赚钱
#
# ''')"""
# lemon = "柠檬"
# print(lemon)
#
# name = "yiyi"
# gender = "female"
# hobby = "赚钱"
# print('''----{}基本信息----
# name: {}
# gender: {}
# hobby: {}
#
# '''.format(name, name, gender, hobby))
# name = "yiyi"
# gender = "female"
# hobby = "赚钱"
# age = "18"
# print('''----%s基本信息----
# name = %s
# gender = %s
# hobby = %s
# age = %s
# ''' % (name, name, gender, hobby, age))
str1 = 'hello lemonban!'
# print(str1[4])
# print(str1[0:5:1])
# print(len(str1))
# print(str1.index('c'))
# print(str1.find('c'))
# print(str1.index('l'))
# print(str1.count('a'))
# print(str1.replace('lemonban','python'))
# num1 = int(input("请输入一个数字："))
# print(type(num1))
# print(num1+2)
# print('abc'-'ab')
# print(2*3)
# print('I love you'*3000)
# print(10/2)
# a = 10
# a = a + 5
# a -=5
# print(a)
# print(4 == 5)
# print('我爱你'=='我爱你')
# print(4>5 and 9>5)
# print(3>5 or  4>8)
# print(not 5<9)
# str2 = "hello python"
# print("p" in str2)
# name = input("请输入姓名：")
# age = input("请输入年龄：")
# # gender = input("请输入性别：")
# # print('''******************
# # 姓名 = {}
# # 年龄 = {}
# # 性别 = {}
# # ******************
# # '''.format(name, age,gender))
# str1 ='python hello aaa 123123aabb'
# # print(str1[7:12:1])
# # print('o a' in str1)
# # print('he' in str1)
# # print('ab' in str1)
# # print(str1.replace('python', 'lemon'))
# # print(str1.index('a'))
# # print(str1.rindex('a '))
#
# list1 = [10, 30, 3.14, True, 'freestyle', [1, 3, 5, 4]]
# # print(list1)
# # print(len(list1))
# # print(list1[4])
# # print(list1[0:5:1])
# # print(list1[5][3])
# # 增加
# list1.append('等')
# print(list1)
# list1.insert(3, '露露')
# print(list1)
# list1.extend(['柠檬', '六六', '天天'])
# print(list1)
# list1.pop(6)
# print(list1)
# list1.remove(10)
# print(list1)
# # ist1.clear()
# # print(list1)
#
# # 修改
# list1[0] = '肥兔'
# print(list1)
# list1[1] = '肥兔'
# print(list1)
# list1.append('肥兔')
# print(list1.count('肥兔'))
#
# tuple1 = ('肥兔', '肥兔', '露露', True, 'freestyle', '等', '柠檬', '六六', '天天')
# print(tuple1)
# # tuple1[0] = '简单'
# # # print(tuple1)
# list2 = list(tuple1)
# print(list2)
# list2[0] = '简单'
# print(list2)
# tuple2 = tuple(list2)
# print(tuple2)


dict1 = {"name": "kki", "age": "18", "gender": "female"}
# print(dict1["age"])
# print(dict1.get("age"))
# print(dict1.keys())
# print(dict1.items())
# print(dict1.values())
# 增加

# dict1["hobby"] = "赚钱"
# print(dict1)
# dict1["gender"] = "male"
# print(dict1)
# dict1.pop("hobby")
# print(dict1)
# print(len(dict1))
# dict1.update({"city": "深圳", "weight": "90", "hight": "165"})
# print(dict1)

#
#
# list3 = [22,11,11,33,44,55]
# set1 = set(list3)
# print(set1)
# lsit4 = list(set1)
# print(lsit4)

# money = int (input("请输入你的余额："))
# if money >= 200:
#     print("做生意！")
# elif money >= 100:
#     print("买房")
# elif money >= 50:
#     print("买车")
# elif money >= 20:
#     print("买零食吃")
# else:
#     print("打工人好好工作，学习！")
# count =0
# # str1 = "我爱学习！"
# # for x in str1:
# #     print(x)
# #     print("*" * 20)
# #     count += 1
# # print(len(str1))
# # print(count)
#
# list1 = ['肥兔', '肥兔', '露露', True, 'freestyle', '等']
# for name in list1:
#     if name == "露露":
#         # break
#         continue
#     print(name)

# tuple2 = ('肥兔', '肥兔', '露露', True, 'freestyle', '等', '柠檬', '六六', '天天')
# for i in tuple2:
#     print(i)
# dict2 = {"name": "kki", "age": "18", "gender": "female"}
# print(len(dict2))
# for h in dict2:
#     print(dict2)
#
# for num in range(0, 19, 3):
#     print(num)


# def good_job():
#     salary = 8000
#     bonus = 2000
#     subsidy = 500
#     sum1 = salary + bonus + subsidy
#     print('工资总和是：{}'.format(sum1))
# good_job()




# def good_job(salary,bonus,subsidy=500,*args,**kwargs):
#     print("参数salary:{}".format(salary))
#     print("参数bonus:{}".format(bonus))
#     print("参数subsidy:{}".format(subsidy))
#     print("参数args:{}".format(args))
#     print("参数kwargs:{}".format(kwargs))
#     sum1 = salary + bonus + subsidy
#     for num in args:
#         sum1 += num
#     for j in kwargs:
#         sum1 += kwargs[j]
#     print('工资总和是：{}'.format(sum1))
# good_job(9000,2000,800,111,222,333,aa=100,bb=200,cc=300)


# def good_job(salary,bonus,subsidy=500,*args,**kwargs):
#     sum1 = salary + bonus + subsidy
#     for num in args:
#         sum1 += num
#     for j in kwargs:
#         sum1 += kwargs[j]
#     return sum1, salary
# result = good_job(9000,2000,800,111,222,333,aa=100,bb=200,cc=300)
# print('工资总和是：{}'.format(result))
# if result[0] >=10000:
#     print("it is a good job")
# else:
#     print("it is not a good job")

# str1 = "我喜欢学习"
# list1 = list(str1)
# print(list1)
# result1 =0
# # for num in range(0,10,2):
# #     result1 += num
# #     print(result1)
# sum(range(0,10,2))
#
# sum1 = 0
# num = range(0, 10, 2)
# for i in num:
#     sum1 += i
#     print(sum1)

# date1 = {"name": "kki", "age": "18", "gender": "female"}
# if len(date1) > 5:
#     print("True")
# else:
#     print("False")

# a = [1, 2, '6', 'summer']
# if 'i' in a:
#     print('i是这个列表的成员！')
# else:
#     print('i不是这个列表的成员！')
#
# dict_1 = {'calss_id': 45, 'num': 20}
# num = dict_1.get('num')
# if num > 5:
#     print("这个班级的人数是：{}".format(num))
# else:
#     print("这个班级的人数不足5人")
#
# list1 = ['肥兔','鹿鹿','Freestyle','等','地球','阑珊','柠檬']
# dict1 ={"name":'肥兔','gender':"female","age":"18",'city':'深圳'}
# dict2 ={"name":'鹿鹿','gender':"female","age":"19",'city':'广州'}
# dict3 ={"name":'Freestyle','gender':"male","age":"20",'city':'中山'}
# dict4 ={"name":'等','gender':"male","age":"19",'city':'东莞'}
# dict5 ={"name":'地球','gender':"male","age":"23",'city':'东莞'}
# dict6 ={"name":'阑珊','gender':"female","age":"18",'city':'南京'}
# dict7 ={"name":'柠檬','gender':"female","age":"18",'city':'南宁'}
# list2 = [dict1,dict2,dict3,dict4,dict5,dict6,dict7]
# print(list2)

# list1 = ['肥兔','鹿鹿','Freestyle','等','地球','阑珊','柠檬']
# list2 = []
# for item in list1:
#     dict1 = dict(name=item, gender="male", age=18, city="北京")
#     list2.append(dict1)
# print(list2)

# list1 = ['肥兔','鹿鹿','Freestyle','等','地球','阑珊','柠檬']
# list2 = ['female','male','male','male','male','female','female']
# list3 = [18,19,20,21,22,19,24]
# list4 = ['深圳','广州','上海','苏州','北京','南京','南宁']
# list5 = []
# for i in range(len(list1)):
#     dict1 = dict(name=list1[i], gender=list2[i], age=list3[i], city=list4[i])
#     list5.append(dict1)
# print(list5)
#
# str1 = "hello,python,lemon,lucy,rita"
# # list1 = list(str1)
# # print(list1)
# list2 = str1.split(',', 2)
# print(list2)

# num1 = int(input('请输入一个数字：'))
# sum1 = 0
# for i in range(num1):
#     sum1 += i
# print(sum1)

# def add_fun(num):
#     sum2 = 0
#     for i in range(num):
#         sum2 += i
#         return sum2
# result = add_fun(50)
# print('这个整数序列相加的和是：{}'.format(result))
#
# def add_fun(num):
#     sum1 = 0
#     for i in range(num):
#         sum1 += i
#     return sum1
# result = add_fun(100)
# print('这个整数序列相加的和是：{}'.format(result))
#
# def object(subject):
#     if type(subject) == list or type(subject) == dict or type(subject) == str:
#         length = len(subject)
#         if length >= 5:
#             print('True')
#         else:
#             print('False')
#     else:
#         print('数据类型不能计算长度')
# object([1,5,12,6,9])



import requests, pprint,jsonpath
# # url_api = "http://8.129.91.152:8766/futureloan/member/register"
# # api_data = {
# #   "mobile_phone": "15015541765",
# #   "pwd": "lemon666",
# #   "type": '1',
# #   "reg_name": "lemon01"
# # }
# # head ={"X-Lemonban-Media-Type": "lemonban.v2",
# # "Content-Type":"application/json"
# # }
# # response = requests.post(url=url_api, json=api_data, headers=head)
# # print(response.json())
#
#
# url_api_login = "http://8.129.91.152:8766/futureloan/member/login"
# api_data_login = {
#   "mobile_phone": "15015541765",
#   "pwd": "lemon666",
# }
# head_login ={"X-Lemonban-Media-Type": "lemonban.v2",
# "Content-Type":"application/json"
# }
# response = requests.post(url=url_api_login, json=api_data_login, headers=head_login)
# pprint.pprint(response.json())
#
# res = response.json()
# '''
# {'code': 0,
#  'copyright': 'Copyright 柠檬班 © 2017-2020 湖南省零檬信息技术有限公司 All Rights Reserved',
#  'data': {'id': 10196100,
#           'leave_amount': 0.0,
#           'mobile_phone': '15015541765',
#           'reg_name': 'lemon01',
#           'reg_time': '2021-01-16 13:51:34.0',
#           'token_info': {'expires_in': '2021-01-16 15:02:28',
#                          'token': 'eyJhbGciOiJIUzUxMiJ9.eyJtZW1iZXJfaWQiOjEwMTk2MTAwLCJleHAiOjE2MTA3ODA1NDh9.jk1xmAIcmlwhzh7JRaZsMgiamDxJaU6ih_irJ3OwtM3fHfUYensiz1MKPSJZfpQuTnOI0ykzQ-jFu6rcfJ-V7Q',
#                          'token_type': 'Bearer'},
#           'type': 1},
#  'msg': 'OK'}
#
# '''
#
# member_id = res["data"]["id"]
# token = res["data"]["token_info"]["token"]
# # print(member_id, token)
# url_api_recharge = "http://8.129.91.152:8766/futureloan/member/recharge"
# api_data_recharge = {
#   "member_id": "member_id",
#   "amount": "5000",
# }
# head_recharge ={"X-Lemonban-Media-Type": "lemonban.v2",
#  "Authorization": "Bearer "+token}
# response = requests.post(url=url_api_recharge, json=api_data_recharge, headers=head_recharge)
# pprint.pprint(response.json())

#注册接口
# url_api_re = "http://8.129.91.152:8766/futureloan/member/register"
# api_data_re = {
#   "mobile_phone": 18835667788,
#   "pwd": "lemon123",
#   "type": "1",
#   "reg_name": "mengmeng"
# }
# head_re = {"X-Lemonban-Media-Type":"lemonban.v2"}
# response = requests.post(url=url_api_re, json=api_data_re, headers=head_re)
# pprint.pprint(response.json())
#
# #登录接口
#
# url_api_login = "http://8.129.91.152:8766/futureloan/member/login"
# api_data_login = {
#   "mobile_phone": 18835667788,
#   "pwd": "lemon123"
# }
# head_login = {"X-Lemonban-Media-Type":"lemonban.v2"}
# response = requests.post(url=url_api_login, json=api_data_login, headers=head_login)
# pprint.pprint(response.json())


#充值接口


#
# res = response.json()
'''
{'code': 0,
 'copyright': 'Copyright 柠檬班 © 2017-2020 湖南省零檬信息技术有限公司 All Rights Reserved',
 'data': {'id': 10196236,
          'leave_amount': 0.0,
          'mobile_phone': '18835667788',
          'reg_name': 'mengmeng',
          'reg_time': '2021-01-16 15:11:37.0',
          'token_info': {'expires_in': '2021-01-16 15:19:44',
                         'token': 'eyJhbGciOiJIUzUxMiJ9.eyJtZW1iZXJfaWQiOjEwMTk2MjM2LCJleHAiOjE2MTA3ODE1ODR9.w1yrxdN5sFdtQ1nz5SgiF_UeBNKtRWRgzpJaYqnFY8L4bkLtwRnMzx8UQHZ0LMcW1oeQCyRjVnnCU1IyE4eeew',
                         'token_type': 'Bearer'},
          'type': 1},
 'msg': 'OK'}

'''
#字典取值   接口关联
# #jsonpath ：安装第三方库，导入  结果放在列表里，要对列表进行取值[]
# res = response.json()
# # member_id = res["data"]["id"]
# # token = res["data"]["token_info"]["token"]
#
# member_id = jsonpath.jsonpath(res, "$..id")[0]
# token = jsonpath.jsonpath(res, "$..token")[0]
# print(member_id, token)
# url_api_rech = "http://8.129.91.152:8766/futureloan/member/recharge"
# api_data_rech = {
#   "member_id": member_id,
#   "amount": 6300
# }
# head_rech = {"X-Lemonban-Media-Type": "lemonban.v2", "Authorization": "Bearer "+token}
# response = requests.post(url=url_api_rech, json=api_data_rech, headers=head_rech)
# pprint.pprint(response.json())


# url_baidu = "https://www.baidu.com/s"
# head = {"User-Agent": "Mozilla/5.0 (Windows NT 6.1; Win64; x64; rv:84.0) Gecko/20100101 Firefox/84.0"}
# param = {"wd": "柠檬班"}
# response = requests.get(url=url_baidu, headers=head, params=param)
# # print(response.text) #自动解码
# print(response.content.decode("utf8"))


import requests, pprint, jsonpath
from openpyxl import load_workbook

#注册接口
def api_func(url_api, data_api):
    head = {"X-Lemonban-Media-Type":"lemonban.v2"}
    response = requests.post(url=url_api, json=data_api, headers=head)
    return response.json()

# url_api = "http://8.129.91.152:8766/futureloan/member/register"
# data_api = {
#       "mobile_phone": 18835667788,
#       "pwd": "lemon123",
#       "type": "1",
#       "reg_name": "mengmeng"
#     }
# resp = api_func(url_api, data_api)
# print(resp)


# 读取数据：第三方库
#读取数据的函数
 #读取数据
def read_data(filename, sheetname):
    wb = load_workbook(filename)     #找表格
    sheet = wb[sheetname]      #找表单
    cases = []
    max_row = sheet.max_row
    for i in range(2, max_row+1):   #遍历表格中行列
        case1 = dict(
        case_id =sheet.cell(row=i, column=1).value,
        url =sheet.cell(row=i, column=5).value,   #获取单元格的值
        data =sheet.cell(row=i, column=6).value,
        expected_result =sheet.cell(row=i, column=7).value)
        cases.append(case1)
    return cases


#写入数据
def write_result(filename, sheetname, final_result, row, column):
    wb = load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row=row, column=column).value = final_result
    wb.save(filename)




